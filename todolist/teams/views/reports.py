# Standard libs
import calendar
from collections import defaultdict
from datetime import datetime, timedelta
import os
import requests

# Django
from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string

# Third-party
import openpyxl
from openpyxl.styles import Font

# Local apps
from main.models import Project
from teams.decorators import employer_required


def seconds_to_hm(seconds):
    if seconds <= 60 * 15:
        return "—"
    
    rounded_hours = round((seconds / 3600) * 2) / 2

    hours = int(rounded_hours)
    minutes = int((rounded_hours - hours) * 60)

    return f"{hours:02d}:{minutes:02d}"

def round_to_half_hour(hours_float):
    return round(hours_float * 2) / 2

def get_weekly_report_html(request, project, start_of_week, end_of_week):
    week_dates = [start_of_week + timedelta(days=i) for i in range(7)]
    company_employees = project.company.employees.all()

    employee_data = []
    totals_by_day = [0] * 7
    grand_total = 0

    for employee in company_employees:
        employee_full_name = employee.get_full_name()
        daily_hours = []
        employee_total = 0

        for i, current_date in enumerate(week_dates):
            entries = employee.time_entries.filter(
                start_time__date=current_date,
                project=project
            )
            total_seconds = sum(entry.duration.total_seconds() for entry in entries)
            employee_total += total_seconds
            totals_by_day[i] += total_seconds
            daily_hours.append(seconds_to_hm(total_seconds))

        if employee_total == 0:
            continue

        employee_data.append({
            "employee_name": employee_full_name,
            "employee_times": daily_hours,
            "employee_total": seconds_to_hm(employee_total)
        })
        grand_total += employee_total

    context = {
        'request': request,
        'project': project,
        'week_dates': week_dates,
        'project_row': [seconds_to_hm(s) for s in totals_by_day],
        'project_total': seconds_to_hm(grand_total),
        'employee_data': employee_data,
        'start_date': start_of_week.strftime('%d/%m/%y'),
        'end_date': end_of_week.strftime('%d/%m/%y'),
    }

    return render_to_string('teams/project_weekly_report.html', context)

@login_required
@employer_required
def project_weekly_report(request, project_id):
    project = get_object_or_404(Project, id=project_id)

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if start_date_str and end_date_str:
        start_of_week = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_of_week = datetime.strptime(end_date_str, '%Y-%m-%d')
    else:
        today = datetime.today()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)

    week_dates = [start_of_week + timedelta(days=i) for i in range(7)]

    company_employees = project.company.employees.all()

    employee_data = []
    totals_by_day = [0] * 7
    grand_total = 0

    for employee in company_employees:
        employee_full_name = employee.get_full_name()
        daily_hours = []
        employee_total = 0

        for i, current_date in enumerate(week_dates):
            entries = employee.time_entries.filter(
                start_time__date=current_date,
                project=project
            )
            total_seconds = sum(entry.duration.total_seconds() for entry in entries)
            employee_total += total_seconds
            totals_by_day[i] += total_seconds
            daily_hours.append(seconds_to_hm(total_seconds))

        if employee_total == 0:
            continue

        employee_data.append({
            "employee_name": employee_full_name,
            "employee_times": daily_hours,
            "employee_total": seconds_to_hm(employee_total)
        })
        grand_total += employee_total

    project_row = [seconds_to_hm(s) for s in totals_by_day]
    project_total = seconds_to_hm(grand_total)

    context = {
        'request': request,
        'project': project,
        'week_dates': week_dates,
        'project_row': project_row,
        'project_total': project_total,
        'employee_data': employee_data,
        'start_date': start_of_week.strftime('%d/%m/%y'),
        'end_date': end_of_week.strftime('%d/%m/%y'),
    }

    html = render_to_string('teams/project_weekly_report.html', context)
    
    response = requests.post(
        "https://html2pdf.fly.dev/api/generate",
        json={
            "html": html,
            "format": "A3",
            "landscape": True,
            "filename": f"week_{start_of_week.isocalendar().week}",
            "printBackground": True
        }
    )

    if response.status_code != 200:
        return JsonResponse({'success': False, 'error': 'PDF conversion failed'}, status=500)

    pdf_bytes = response.content

    filename = f"TimeReport_{project.title}_Week-{start_of_week.isocalendar().week}_{start_of_week.strftime('%d.%m.%Y')}-{end_of_week.strftime('%d.%m.%Y')}.pdf"
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
@employer_required
def project_monthly_report_pdf(request, project_id):
    project = get_object_or_404(Project, id=project_id)

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if start_date_str and end_date_str:
        start_of_month = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_of_month = datetime.strptime(end_date_str, '%Y-%m-%d')
    else:
        today = datetime.today()
        start_of_month = today.replace(day=1)
        _, last_day = calendar.monthrange(today.year, today.month)
        end_of_month = today.replace(day=last_day)

    current = start_of_month
    file_links = []

    while current <= end_of_month:
        start_of_week = current
        end_of_week = min(current + timedelta(days=6), end_of_month)

        html = get_weekly_report_html(request, project, start_of_week, end_of_week)

        response = requests.post(
            "https://html2pdf.fly.dev/api/generate",
            json={
                "html": html,
                "format": "A3",
                "landscape": True,
                "filename": f"week_{start_of_week.isocalendar().week}",
                "printBackground": True
            }
        )

        if response.status_code != 200:
            return JsonResponse({'success': False, 'error': 'PDF conversion failed'}, status=500)
        
        filename = f"TimeReport_{project.title}_Week-{start_of_week.isocalendar().week}_{start_of_week.strftime('%d.%m.%Y')}-{end_of_week.strftime('%d.%m.%Y')}.pdf"
        path = os.path.join('tmp', 'reports', filename)

        default_storage.save(path, ContentFile(response.content))

        file_url = default_storage.url(path)
        file_links.append(file_url)

        current += timedelta(days=7)
    
    file_links.sort()
    return JsonResponse({'success': True, 'links': file_links}, status=200)

@login_required
@employer_required
def project_monthly_report_xlsx(request, project_id):
    project = get_object_or_404(Project, id=project_id)

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        start_of_month = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_of_month = datetime.strptime(end_date_str, '%Y-%m-%d')
    else:
        today = datetime.today()
        start_of_month = today.replace(day=1)
        _, last_day = calendar.monthrange(today.year, today.month)
        end_of_month = today.replace(day=last_day)

    month_days = (end_of_month - start_of_month).days + 1
    month_dates = [start_of_month + timedelta(days=i) for i in range(month_days)]

    week_to_dates = defaultdict(list)
    for date in month_dates:
        week_num = date.isocalendar().week
        week_to_dates[week_num].append(date)

    job_week_data = defaultdict(float)

    for employee in project.company.employees.all():
        job_title = employee.job_title or "Unknown"
        for week_num, dates in week_to_dates.items():
            total_seconds = 0
            for date in dates:
                entries = employee.time_entries.filter(start_time__date=date, project=project)
                total_seconds += sum(entry.duration.total_seconds() for entry in entries)
            if total_seconds > 0:
                job_week_data[(job_title, week_num)] += total_seconds

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    ws.append(["Job Title - Project | Week", "Total Time"])
    ws["A1"].font = ws["B1"].font = Font(bold=True)

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 30

    for (job_title, week_num), total_seconds in sorted(
        job_week_data.items(), 
        key=lambda x: (x[0][1], x[0][0].name if hasattr(x[0][0], 'name') else str(x[0][0]))
    ):
        hours_float = total_seconds / 3600
        rounded_hours = round_to_half_hour(hours_float)
        formatted_time = f"{rounded_hours:.1f}"
        label = f"{job_title} - {project.title} | week {week_num}"
        ws.append([label, formatted_time])

    ws.append([""])
    ws.append(["Period", f"{start_of_month.strftime('%d/%m/%Y')} - {end_of_month.strftime('%d/%m/%Y')}"])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"TimeReport_{project.title}_{start_of_month.strftime('%d.%m.%Y')}-{end_of_month.strftime('%d.%m.%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
